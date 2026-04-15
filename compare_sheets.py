import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import glob

# Config
HEADER_ROW = 5
START_ROW = 6
KEY_COL_IDX = 9  # Column J (0-indexed)

# Colors
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # Added (in N, not in O)
GREY_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')     # Removed (in O, not in N)
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')   # Value changed
GREEN_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')    # New column in N
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')


def extract_rows(sheet):
    """Return {key: (row_idx, [cell_values])} for data rows."""
    data = {}
    for r_idx, row in enumerate(sheet.iter_rows(min_row=START_ROW), start=START_ROW):
        key = row[KEY_COL_IDX].value
        if key is None:
            continue
        data[str(key).strip()] = (r_idx, [c.value for c in row])
    return data


def fill_row(sheet, row_idx, num_cols, fill):
    for c_idx in range(1, num_cols + 1):
        sheet.cell(row=row_idx, column=c_idx).fill = fill


def values_equal(a, b):
    sa = "" if a is None else str(a).strip()
    sb = "" if b is None else str(b).strip()
    if sa == sb:
        return True
    try:
        return float(sa) == float(sb)
    except (ValueError, TypeError):
        return False


def compare_sheets(file_path, output_dir):
    print(f"Loading {file_path}...")
    wb_vals = openpyxl.load_workbook(file_path, data_only=True)
    wb_styles = openpyxl.load_workbook(file_path)

    if 'N' not in wb_vals.sheetnames or 'O' not in wb_vals.sheetnames:
        print("Error: Sheet 'N' or 'O' not found in the workbook.")
        return

    n_vals, o_vals = wb_vals['N'], wb_vals['O']
    n_styles, o_styles = wb_styles['N'], wb_styles['O']

    n_headers = [c.value for c in n_vals[HEADER_ROW]]
    o_headers = [c.value for c in o_vals[HEADER_ROW]]
    o_header_set = {str(h).strip() for h in o_headers if h is not None}

    n_data = extract_rows(n_vals)
    o_data = extract_rows(o_vals)

    diff_list = []

    # 1. New columns in N -> Green
    print("Highlighting new columns (Green)...")
    for c_idx, h in enumerate(n_headers, start=1):
        if h and str(h).strip() not in o_header_set:
            n_styles.cell(row=HEADER_ROW, column=c_idx).fill = GREEN_FILL

    # 2. Added rows (in N, not in O) -> Yellow
    print("Highlighting added rows (Yellow)...")
    for key in n_data.keys() - o_data.keys():
        r_idx, _ = n_data[key]
        fill_row(n_styles, r_idx, len(n_headers), YELLOW_FILL)
        diff_list.append(['Added', key, 'Whole Row', '(Missing in O)', 'Present in N'])

    # 3. Removed rows (in O, not in N) -> Grey
    print("Highlighting removed rows (Grey)...")
    for key in o_data.keys() - n_data.keys():
        r_idx, _ = o_data[key]
        fill_row(o_styles, r_idx, len(o_headers), GREY_FILL)
        diff_list.append(['Removed', key, 'Whole Row', 'Present in O', '(Missing in N)'])

    # 4. Changed values -> Orange
    print("Highlighting changed values (Orange)...")
    for key in n_data.keys() & o_data.keys():
        r_idx_n, vals_n = n_data[key]
        _, vals_o = o_data[key]
        for c_idx in range(KEY_COL_IDX + 1, len(vals_n)):
            v_n = vals_n[c_idx]
            v_o = vals_o[c_idx] if c_idx < len(vals_o) else None
            if values_equal(v_n, v_o):
                continue
            n_styles.cell(row=r_idx_n, column=c_idx + 1).fill = ORANGE_FILL
            so = "" if v_o is None else str(v_o).strip()
            sn = "" if v_n is None else str(v_n).strip()
            diff_list.append(['Changed', key, str(n_headers[c_idx]), so, sn])

    # 5. Summary sheet
    print("Writing Summary sheet...")
    if 'Summary' in wb_styles.sheetnames:
        del wb_styles['Summary']
    ws_sum = wb_styles.create_sheet('Summary', 0)

    headers = ['Type', 'Line No', 'Field', 'Old Value (O)', 'New Value (N)']
    for i, h in enumerate(headers, start=1):
        c = ws_sum.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center')

    for i, entry in enumerate(diff_list, start=2):
        for j, val in enumerate(entry, start=1):
            ws_sum.cell(row=i, column=j, value=val)

    total_row = len(diff_list) + 3
    ws_sum.cell(row=total_row, column=1, value='Total Changes:').font = Font(bold=True)
    ws_sum.cell(row=total_row, column=2, value=len(diff_list)).font = Font(bold=True)

    for i in range(1, len(headers) + 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = 25

    # 6. Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(output_dir, f"compare_result_{timestamp}.xlsx")
    wb_styles.save(output_path)

    added = len(n_data.keys() - o_data.keys())
    removed = len(o_data.keys() - n_data.keys())
    changed = len(diff_list) - added - removed
    print(f"Success! Found {added} Added, {removed} Removed, {changed} Value Changes.")
    print(f"Results saved to: {output_path}")


if __name__ == "__main__":
    input_dir = r'C:\Users\dc13691\Desktop\Compare\Input'
    output_dir = r'C:\Users\dc13691\Desktop\Compare\Output'

    os.makedirs(output_dir, exist_ok=True)
    files = glob.glob(os.path.join(input_dir, '*.xlsx'))
    if not files:
        print(f"No .xlsx files found in {input_dir}")
    else:
        latest = max(files, key=os.path.getmtime)
        compare_sheets(latest, output_dir)
